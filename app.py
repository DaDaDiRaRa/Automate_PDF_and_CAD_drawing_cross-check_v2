"""
app.py  —  PDF <-> CAD Drawing Cross-Check  (V46)
================================================
V46 변경사항:
* 도면명 절삭(Truncation) 버그 완벽 해결: "상세도-2", "마감표-1" 등의 도면명 끝부분이 
  도면번호 정규식 패턴(글자+숫자)과 일치하여 도면명에서 잘려나가는 문제를 수정했습니다.
* 지능형 키워드 필터 도입: 도면번호 후보를 찾았을 때, 그 앞글자(Prefix)가 '도' 또는 '표'로 
  끝나거나 도면명 전용 키워드(상세, 일람 등)를 포함하면 도면번호로 인식하지 않고 패스합니다.
"""

from __future__ import annotations
import glob, os, re, traceback
from pathlib import Path
from typing import List, Optional, Tuple

import pandas as pd
import pdfplumber
import ezdxf
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# ============================================================================
# 전역 설정
# ============================================================================
ODA_PATH: str = ""
리포트_이름: str = "도면검토리포트_V46.xlsx"
DEBUG: bool = True  

X_MIN_RATIO: float = 0.850
X_MAX_RATIO: float = 1.050 
Y_MIN_RATIO: float = -0.050 
Y_MAX_RATIO: float = 0.250 

_도면번호_패턴 = re.compile(r"([A-Z\u0391-\u03A9\.가-힣]{1,4})[-_ ]*(\d{1,5}[A-Z]*|TOE)")
_축척_패턴 = re.compile(r"(1\s?[/:,]\s?(\d{1,4})|NONE|N/A)", re.I)

def _도면번호_세척(raw_s: str) -> str:
    if not raw_s: return ""
    s = raw_s.strip().upper().replace("Λ", "A").replace("Δ", "A").replace("TOE", "108")
    if s.startswith("."): s = "AA" + s[1:]
    return re.sub(r"\s+", " ", s)

def _축척_텍스트_정리(txt: str) -> str:
    if not txt: return "X"
    upper_txt = txt.upper()
    if "NONE" in upper_txt or "N/A" in upper_txt: return "NONE"
    m = _축척_패턴.search(upper_txt)
    if m and m.group(2): return f"1/{m.group(2)}"
    return "X"

# V46 핵심: 도면명이 도면번호로 오인되는 것을 막는 지능형 추출 함수
def _extract_drawing_number(text: str) -> Optional[str]:
    for m in _도면번호_패턴.finditer(text):
        prefix = m.group(1)
        # 앞글자가 '도'(상세도)나 '표'(일람표)로 끝나면 도면번호가 아님!
        if prefix.endswith("도") or prefix.endswith("표"): continue
        if any(k in prefix for k in ["상세", "일람", "배치", "전개", "마감", "계획", "조감", "구조"]): continue
        return m.group(0) # 유효한 도면번호만 반환
    return None

# ============================================================================
# 1. PDF 데이터 추출
# ============================================================================
def extract_pdf_table(pdf_path: str) -> pd.DataFrame:
    print(f"[PDF ] 분석 시작: {os.path.basename(pdf_path)}")
    데이터 = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                width, height = page.width, page.height
                경계선 = width * 0.465
                for 범위 in [(0, 0, mid := width * 0.465, height), (mid, 0, width, height)]:
                    영역 = page.crop(범위)
                    텍스트 = 영역.extract_text()
                    if not 텍스트: continue
                    for 줄 in 텍스트.splitlines():
                        줄 = re.sub(r"(\d)(1\s?[/:,])", r"\1 \2", 줄)
                        
                        # V46 안전한 도면번호 추출 적용
                        raw_no = _extract_drawing_number(줄)
                        if not raw_no: continue
                        번호 = _도면번호_세척(raw_no)
                        남은텍스트 = 줄.replace(raw_no, "").strip()
                        
                        found_scales = []
                        for token in 남은텍스트.split():
                            if "NONE" in token.upper() or "N/A" in token.upper():
                                found_scales.append("NONE")
                            else:
                                sm = _축척_패턴.search(token)
                                if sm: found_scales.append(f"1/{sm.group(2)}")
                                
                        a1 = found_scales[0] if len(found_scales) >= 1 else "X"
                        a3 = found_scales[1] if len(found_scales) >= 2 else "X"
                        
                        명칭 = 남은텍스트
                        for s in found_scales:
                            if s != "NONE": 명칭 = re.sub(rf"1\s?[/:,]\s?{s[2:]}", "", 명칭)
                        명칭 = 명칭.replace("NONE", "").replace("N/A", "").strip().strip(",").strip()
                        if 번호:
                            데이터.append({"도면번호(PDF)": 번호, "도면명(PDF)": 명칭, "축척_A1(PDF)": a1, "축척_A3(PDF)": a3})
                print(f"[PDF ] {page.page_number}페이지 분석 완료")
    except Exception as e: print(f"[ERROR] PDF 읽기 실패: {e}")
    df = pd.DataFrame(데이터)
    return df.drop_duplicates(subset=["도면번호(PDF)"]).reset_index(drop=True) if not df.empty else df

# ============================================================================
# 2. CAD 데이터 추출
# ============================================================================
def _oda_환경_설정():
    from ezdxf.addons import odafc
    설치경로 = ""
    for 경로 in [r"C:\Program Files\ODA", r"C:\Program Files (x86)\ODA"]:
        실행파일들 = glob.glob(os.path.join(경로, "*", "ODAFileConverter.exe"))
        if 실행파일들: 설치경로 = sorted(실행파일들, reverse=True)[0]; break
    if 설치경로:
        폴더경로 = os.path.dirname(설치경로)
        if 폴더경로 not in os.environ.get("PATH", ""):
            os.environ["PATH"] = 폴더경로 + os.pathsep + os.environ.get("PATH", "")
        try: ezdxf.options.set("odafc", "win_exec_path", 설치경로)
        except: pass
    return 설치경로

def _cad_문서_로드(path: Path):
    확장자 = path.suffix.lower()
    if 확장자 == ".dxf": return ezdxf.readfile(str(path))
    _oda_환경_설정()
    from ezdxf.addons import odafc
    return odafc.readfile(str(path))

def _텍스트_데이터_추출(ent) -> Optional[Tuple[float, float, str]]:
    유형 = ent.dxftype()
    try:
        if 유형 == "TEXT":
            p = ent.dxf.align_point if getattr(ent.dxf, "halign", 0) or getattr(ent.dxf, "valign", 0) else ent.dxf.insert
            return (float(p[0]), float(p[1]), (ent.dxf.text or "").strip())
        if 유형 == "MTEXT":
            return (float(ent.dxf.insert[0]), float(ent.dxf.insert[1]), ent.plain_text().strip())
    except Exception: return None
    return None

def extract_dwg_data(target_dir: str, block_name: str, base_w: float, base_h: float) -> pd.DataFrame:
    목표블록 = block_name.strip().lower()
    데이터 = []
    캐드파일들 = sorted(set(glob.glob(os.path.join(target_dir, "*.dwg")) + glob.glob(os.path.join(target_dir, "*.dxf"))))

    for i, 전체경로 in enumerate(캐드파일들, 1):
        파일명 = os.path.basename(전체경로)
        print(f"[CAD ] ({i}/{len(캐드파일들)}) {파일명:<30}", end=" | ", flush=True)
        try:
            doc = _cad_문서_로드(Path(전체경로))
            msp = doc.modelspace()
            도곽들 = [ins for ins in msp.query("INSERT") if 목표블록 in ins.dxf.name.lower()]
            if not 도곽들: print("도곽 없음."); continue

            모든텍스트 = []
            for ent in msp.query("TEXT MTEXT"):
                d = _텍스트_데이터_추출(ent)
                if d and d[2]: 모든텍스트.append(d)

            추출개수 = 0
            for idx, 도곽 in enumerate(도곽들, 1):
                ix, iy = 도곽.dxf.insert.x, 도곽.dxf.insert.y
                xscale, yscale = abs(도곽.dxf.xscale), abs(도곽.dxf.yscale)
                너비, 높이 = base_w * xscale, base_h * yscale
                x_min, x_max = ix + (너비 * X_MIN_RATIO), ix + (너비 * X_MAX_RATIO)
                y_min, y_max = iy + (높이 * Y_MIN_RATIO), iy + (높이 * Y_MAX_RATIO)

                박스텍스트 = [t for t in 모든텍스트 if x_min <= t[0] <= x_max and y_min <= t[1] <= y_max]
                if 박스텍스트:
                    
                    labels_a1 = [t for t in 박스텍스트 if re.search(r"\bA1\b", t[2].upper())]
                    labels_a3 = [t for t in 박스텍스트 if re.search(r"\bA3\b", t[2].upper())]
                    scales = [t for t in 박스텍스트 if _축척_패턴.search(t[2].upper())]
                    
                    a1, a3 = "X", "X"
                    
                    if labels_a1 and scales:
                        a1_y = labels_a1[0][1]
                        closest_to_a1 = min(scales, key=lambda s: abs(s[1] - a1_y))
                        if abs(closest_to_a1[1] - a1_y) < (높이 * 0.05):
                            a1 = _축척_텍스트_정리(closest_to_a1[2])
                            scales.remove(closest_to_a1)

                    if labels_a3 and scales:
                        a3_y = labels_a3[0][1]
                        closest_to_a3 = min(scales, key=lambda s: abs(s[1] - a3_y))
                        if abs(closest_to_a3[1] - a3_y) < (높이 * 0.05):
                            a3 = _축척_텍스트_정리(closest_to_a3[2])
                            
                    if a1 == "X" and a3 == "X" and scales:
                        scales.sort(key=lambda x: -x[1])
                        if len(scales) >= 1: a1 = _축척_텍스트_정리(scales[0][2])
                        if len(scales) >= 2: a3 = _축척_텍스트_정리(scales[1][2])

                    박스텍스트.sort(key=lambda x: -x[1])
                    줄목록 = []
                    현재_줄, 현재_y = [], None
                    for tx, ty, txt in 박스텍스트:
                        if 현재_y is None: 현재_y = ty; 현재_줄.append((tx, txt))
                        elif abs(현재_y - ty) < (높이 * 0.015): 현재_줄.append((tx, txt))
                        else:
                            현재_줄.sort(key=lambda x: x[0]); 줄목록.append(" ".join([x[1] for x in 현재_줄]))
                            현재_y = ty; 현재_줄 = [(tx, txt)]
                    if 현재_줄: 현재_줄.sort(key=lambda x: x[0]); 줄목록.append(" ".join([x[1] for x in 현재_줄]))

                    번호, 명칭 = "", ""
                    명칭후보 = []
                    for 줄 in 줄목록:
                        # V46 안전한 도면번호 추출 적용
                        raw_no = _extract_drawing_number(줄)
                        if raw_no and not 번호: 번호 = _도면번호_세척(raw_no)
                        
                        clean = 줄.replace(raw_no if raw_no else "", "").strip()
                        clean = re.sub(r"\bA1\b|\bA3\b|NONE|N/A", "", clean, flags=re.IGNORECASE)
                        clean = re.sub(r"1\s?[/:,]\s?\d{1,4}", "", clean, flags=re.IGNORECASE)
                        clean = clean.strip().strip(",")
                        
                        if len(clean) > 2 and "도면명" not in clean and "SCALE" not in clean.upper():
                            명칭후보.append(clean)
                    if 명칭후보: 명칭 = max(명칭후보, key=lambda s: len(s.replace(" ", "")))

                    if DEBUG: print(f"    [DBG] 도곽 #{idx} 확정 결과: A1={a1}, A3={a3}, 번호={번호}, 명칭={명칭}")
                    if 번호:
                        데이터.append({"파일명": 파일명, "도면번호(DWG)": 번호, "도면명(DWG)": 명칭, "축척_A1(DWG)": a1, "축척_A3(DWG)": a3})
                        추출개수 += 1
            print(f"성공 ({추출개수}/{len(도곽들)})")
        except Exception as e: print(f"오류: {e}")
    return pd.DataFrame(데이터) if 데이터 else pd.DataFrame(columns=["파일명", "도면번호(DWG)", "도면명(DWG)", "축척_A1(DWG)", "축척_A3(DWG)"])

# ============================================================================
# 3. 리포트 생성 및 메인
# ============================================================================
def build_report(pdf_df: pd.DataFrame, dwg_df: pd.DataFrame, out_path: str):
    pdf, dwg = pdf_df.copy(), dwg_df.copy()
    pdf["KEY"] = pdf["도면번호(PDF)"].str.replace(" ", "") if "도면번호(PDF)" in pdf.columns else ""
    dwg["KEY"] = dwg["도면번호(DWG)"].str.replace(" ", "") if "도면번호(DWG)" in dwg.columns else ""
    결과 = pd.merge(pdf, dwg, on="KEY", how="outer", indicator=True)
    결과["상태"] = 결과["_merge"].map({"both": "일치", "left_only": "DWG 누락", "right_only": "PDF 누락"})
    정리할_컬럼 = ["도면번호(PDF)", "도면명(PDF)", "축척_A1(PDF)", "축척_A3(PDF)", "도면번호(DWG)", "도면명(DWG)", "축척_A1(DWG)", "축척_A3(DWG)", "파일명", "상태"]
    for c in 정리할_컬럼:
        if c not in 결과.columns: 결과[c] = ""
    결과 = 결과[정리할_컬럼].fillna("X")
    결과.to_excel(out_path, index=False)
    
    빨간색 = PatternFill(start_color="FFFF9999", end_color="FFFF9999", fill_type="solid")
    wb = load_workbook(out_path); ws = wb.active
    h = {cell.value: cell.column for cell in ws[1]}
    for row in range(2, ws.max_row + 1):
        상태 = ws.cell(row=row, column=h.get("상태", 1)).value
        if 상태 != "일치":
            for c in h.values(): ws.cell(row=row, column=c).fill = 빨간색
            continue
        if str(ws.cell(row=row, column=h.get("도면번호(PDF)")).value).replace(" ","") != str(ws.cell(row=row, column=h.get("도면번호(DWG)")).value).replace(" ",""):
            ws.cell(row=row, column=h.get("도면번호(PDF)")).fill = 빨간색; ws.cell(row=row, column=h.get("도면번호(DWG)")).fill = 빨간색
        for s in ["A1", "A3"]:
            p_val = str(ws.cell(row=row, column=h.get(f"축척_{s}(PDF)")).value).replace(" ","")
            d_val = str(ws.cell(row=row, column=h.get(f"축척_{s}(DWG)")).value).replace(" ","")
            if p_val != d_val:
                ws.cell(row=row, column=h.get(f"축척_{s}(PDF)")).fill = 빨간색; ws.cell(row=row, column=h.get(f"축척_{s}(DWG)")).fill = 빨간색
    wb.save(out_path)
    print(f"[XLSX] 리포트 저장 완료: {out_path}")

def main():
    print("=" * 72); print(" PDF <-> CAD 체크 V46 (도면명 절삭 방지 및 축척 완벽 적용)"); print("=" * 72)
    캐드경로 = input("1. DWG 폴더 경로: ").strip().strip('"')
    PDF경로 = input("2. PDF 파일 경로: ").strip().strip('"')
    블록이름 = input("3. 도곽 블록 이름: ").strip()
    try:
        base_w = float(input("4. 해당 도곽 원본의 가로 길이를 입력하세요 (예: 841): ").strip())
        base_h = float(input("5. 해당 도곽 원본의 세로 길이를 입력하세요 (예: 594): ").strip())
    except: base_w, base_h = 841.0, 594.0
    pdf_데이터 = extract_pdf_table(PDF경로)
    dwg_데이터 = extract_dwg_data(캐드경로, 블록이름, base_w, base_h)
    build_report(pdf_데이터, dwg_데이터, os.path.abspath(리포트_이름))
    print("-" * 72); print("[DONE] 작업이 완료되었습니다.")

if __name__ == "__main__":
    main()
